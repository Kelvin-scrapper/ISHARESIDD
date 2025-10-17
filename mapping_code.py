"""
Robust Mapping Script - ETF Effective Duration Data
- DUAL MAPPING LOGIC:
  1. First checks and updates/corrects HISTORICAL data (from "As of Date" in source)
  2. Then maps to LAST BUSINESS DATE (skips weekends)
- WEEKEND HANDLING:
  * Monday -> Maps to Friday (skips weekend)
  * Saturday -> Maps to Friday
  * Sunday -> Maps to Friday
  * Tuesday-Friday -> Maps to previous day
- Handles 0 values by using previous day's data
- Reviews and corrects historical data if changes detected
- Validates data before writing
- Collaborates with main.py for data extraction
"""

import openpyxl
from datetime import datetime, timedelta
import os

# HARDCODED HEADERS - Row 1 (Technical Identifiers)
HEADER_ROW_1 = [
    None,  # Column A (reserved for date)
    'ISHARESIDD.JPMEMBONDSETF.EFFECTDUR.B',
    'ISHARESIDD.IBOXXHIGHYIELDCORPBONDETF.EFFECTDUR.B',
    'ISHARESIDD.IBOXXIGCORPBONDETF.EFFECTDUR.B',
    'ISHARESIDD.TIPSBONDETF.EFFECTDUR.B',
    'ISHARESIDD.20YRTREASURYBONDETF.EFFECTDUR.B',
    'ISHARESIDD.710YEARTREASURYBONDETF.EFFECTDUR.B',
    'ISHARESIDD.JPMORGANEMBONDUCITSETFUSD.EFFECTDUR.B',
    'ISHARESIDD.TREASURYBOND13YRUCITSETFUSD.EFFECTDUR.B',
    'ISHARESIDD.HIGHYIELDCORPBONDUCITSETFUSD.EFFECTDUR.B',
    'ISHARESIDD.CORPBOND15YRUCITSETFEUR.EFFECTDUR.B',
    'ISHARESIDD.HIGHYIELDCORPBONDUCITSETFEUR.EFFECTDUR.B',
    'ISHARESIDD.CORECORPBONDUCITSETFEUR.EFFECTDUR.B',
    'ISHARESIDD.MBSETF.EFFECTDUR.B',
    'ISHARESIDD.SHORTTERMCORPORATEBONDETF.EFFECTDUR.B',
    'ISHARESIDD.INTERMEDIATETERMCORPORATEBONDETF.EFFECTDUR.B'
]

# HARDCODED HEADERS - Row 2 (ETF Names)
HEADER_ROW_2 = [
    None,  # Column A (reserved for date)
    'JPM EM Bonds ETF',
    'Iboxx High Yield Corp Bond ETF',
    'Iboxx IG Corp Bond ETF',
    'TIPS Bond ETF',
    '20 Yr Treasury Bond ETF',
    '7-10 Year Treasury Bond ETF',
    'iShares J.P. Morgan $ EM Bond UCITS ETF',
    'iShares $ Treasury Bond 1-3yr UCITS ETF',
    'iShares $ High Yield Corp Bond UCITS ETF',
    'iShares € Corp Bond 1-5yr UCITS ETF',
    'iShares € High Yield Corp Bond UCITS ETF',
    'iShares Core € Corp Bond UCITS ETF',
    'iShares MBS ETF',
    'iShares Short-Term Corporate Bond ETF',
    'iShares Intermediate-Term Corporate Bond ETF'
]

def validate_duration(value):
    """Validate that duration value is numeric and not zero"""
    if value is None:
        return False
    try:
        float_val = float(str(value).replace(' yrs', '').strip())
        # Return False if value is 0
        if float_val == 0:
            return False
        return True
    except:
        return False

def find_row_by_date(ws, target_date_str):
    """
    Find row with matching date
    Returns row number if found, None otherwise
    """
    for row_idx in range(3, ws.max_row + 1):
        cell_date = ws.cell(row=row_idx, column=1).value
        if cell_date:
            if isinstance(cell_date, datetime):
                if cell_date.strftime("%Y-%m-%d") == target_date_str:
                    return row_idx
            elif isinstance(cell_date, str):
                if cell_date == target_date_str:
                    return row_idx
    return None

def get_previous_row_data(ws, current_row):
    """
    Get data from the previous row (previous date)
    Returns dictionary with column index as key and value
    """
    if current_row <= 3:
        return {}

    previous_row = current_row - 1
    previous_data = {}

    for col_idx in range(2, len(HEADER_ROW_2) + 1):
        value = ws.cell(row=previous_row, column=col_idx).value
        if value is not None:
            previous_data[col_idx] = value

    return previous_data

def get_as_of_date_from_source(source_file='ETF_Effective_Duration.xlsx'):
    """
    Read the 'As of Date' from the source file
    Returns the date string
    """
    if not os.path.exists(source_file):
        return None

    wb = openpyxl.load_workbook(source_file)
    ws = wb.active

    # Read the first ETF's "As of Date"
    as_of_date = ws.cell(row=2, column=3).value
    wb.close()

    if isinstance(as_of_date, datetime):
        return as_of_date.strftime("%Y-%m-%d")
    elif isinstance(as_of_date, str):
        return as_of_date

    return None

def read_source_data(source_file='ETF_Effective_Duration.xlsx'):
    """
    Read ETF data from source file
    Returns dictionary with ETF name as key
    """
    if not os.path.exists(source_file):
        print(f"ERROR: {source_file} not found!")
        return None

    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb.active

    source_data = {}

    for row_idx in range(2, source_ws.max_row + 1):  # Skip header row
        etf_name = source_ws.cell(row=row_idx, column=1).value
        effective_duration = source_ws.cell(row=row_idx, column=2).value
        as_of_date = source_ws.cell(row=row_idx, column=3).value

        if etf_name:
            # Clean the effective duration value (remove 'yrs' if present)
            if effective_duration and isinstance(effective_duration, str):
                effective_duration = effective_duration.replace(' yrs', '').strip()

            # Store raw value (validation happens later)
            source_data[etf_name] = {
                'duration': effective_duration,
                'as_of_date': as_of_date
            }

    source_wb.close()
    return source_data

def create_or_open_target_file(output_filename='ISHARESIDD_DATA_.xlsx'):
    """
    Create new file or open existing file
    Returns workbook and worksheet
    """
    file_exists = os.path.exists(output_filename)

    if file_exists:
        target_wb = openpyxl.load_workbook(output_filename)
        target_ws = target_wb.active
    else:
        target_wb = openpyxl.Workbook()
        target_ws = target_wb.active

        # Write hardcoded headers
        # Write Row 1 (Technical Identifiers)
        for col_idx, header_value in enumerate(HEADER_ROW_1, start=1):
            target_ws.cell(row=1, column=col_idx, value=header_value)

        # Write Row 2 (ETF Names)
        for col_idx, header_value in enumerate(HEADER_ROW_2, start=1):
            target_ws.cell(row=2, column=col_idx, value=header_value)

    return target_wb, target_ws

def compare_and_report_changes(ws, row_num, column_mapping, label=""):
    """
    Compare new data with existing data in the row
    Report changes detected
    """
    print(f"\n  Comparing with existing data{' for ' + label if label else ''}:")
    changes_detected = False

    for col_idx, data in column_mapping.items():
        old_value = ws.cell(row=row_num, column=col_idx).value
        new_value = data['duration']

        # Convert to float for comparison
        try:
            old_val_float = float(old_value) if old_value else None
            new_val_float = float(new_value) if new_value else None

            if old_val_float != new_val_float:
                etf_name = HEADER_ROW_2[col_idx - 1]
                print(f"    [CHANGED] {etf_name}: {old_value} -> {new_value}")
                changes_detected = True
        except:
            # If conversion fails, just compare as strings
            if str(old_value) != str(new_value):
                etf_name = HEADER_ROW_2[col_idx - 1]
                print(f"    [CHANGED] {etf_name}: {old_value} -> {new_value}")
                changes_detected = True

    if not changes_detected:
        print("    No changes detected - data is identical")

    return changes_detected

def create_column_mapping(source_data, previous_data, label=""):
    """
    Create column mapping from source data
    Handle zero values using previous data
    """
    column_mapping = {}
    zero_value_count = 0

    for col_idx in range(2, len(HEADER_ROW_2) + 1):  # Start from column 2 (B)
        target_etf_name = HEADER_ROW_2[col_idx - 1]  # Adjust for 0-based index

        if target_etf_name and target_etf_name in source_data:
            duration = source_data[target_etf_name]['duration']

            # Check if duration is 0 or invalid
            if not validate_duration(duration):
                if col_idx in previous_data:
                    # Use previous day's data
                    print(f"    [ZERO/INVALID] {target_etf_name}: Using previous value {previous_data[col_idx]}")
                    column_mapping[col_idx] = {
                        'duration': previous_data[col_idx],
                        'as_of_date': source_data[target_etf_name]['as_of_date']
                    }
                    zero_value_count += 1
                else:
                    print(f"    [WARNING] {target_etf_name}: Invalid value ({duration}) and no previous data")
                    column_mapping[col_idx] = source_data[target_etf_name]
            else:
                column_mapping[col_idx] = source_data[target_etf_name]

    return column_mapping, zero_value_count

def write_data_to_row(ws, row_num, date_str, column_mapping, label=""):
    """
    Write data to specified row
    """
    # Write date in column A
    ws.cell(row=row_num, column=1, value=date_str)

    # Write effective duration values
    for col_idx, data in column_mapping.items():
        ws.cell(row=row_num, column=col_idx, value=data['duration'])

def get_last_business_date():
    """
    Calculate the last business date (skips weekends)
    - If today is Monday (0), return Friday (3 days ago)
    - If today is Sunday (6), return Friday (2 days ago)
    - If today is Saturday (5), return Friday (1 day ago)
    - Otherwise, return yesterday
    """
    today = datetime.now()
    weekday = today.weekday()  # Monday=0, Sunday=6

    if weekday == 0:  # Monday
        last_business_date = today - timedelta(days=3)  # Friday
        print(f"[INFO] Today is Monday - mapping to Friday (last business day)")
    elif weekday == 6:  # Sunday
        last_business_date = today - timedelta(days=2)  # Friday
        print(f"[INFO] Today is Sunday - mapping to Friday (last business day)")
    elif weekday == 5:  # Saturday
        last_business_date = today - timedelta(days=1)  # Friday
        print(f"[INFO] Today is Saturday - mapping to Friday (last business day)")
    else:  # Tuesday-Friday
        last_business_date = today - timedelta(days=1)  # Yesterday

    return last_business_date

def create_mapping():
    """
    DUAL MAPPING LOGIC:
    1. Check and update HISTORICAL data (from "As of Date")
    2. Map to LAST BUSINESS DATE (skips weekends)
    """
    output_filename = 'ISHARESIDD_DATA_.xlsx'

    # Calculate LAST BUSINESS DATE (skip weekends)
    today = datetime.now()
    last_business_date = get_last_business_date()
    target_date_str = last_business_date.strftime("%Y-%m-%d")

    print("=" * 80)
    print("ROBUST ETF DATA MAPPING - DUAL DATE LOGIC (BUSINESS DAYS)")
    print("=" * 80)
    print(f"Today's Date: {today.strftime('%Y-%m-%d')} ({today.strftime('%A')})")
    print(f"Target Date: {target_date_str} ({last_business_date.strftime('%A')})")
    print("=" * 80)

    # Get "As of Date" from source file
    as_of_date_str = get_as_of_date_from_source()
    if not as_of_date_str:
        print("ERROR: Could not read 'As of Date' from source file!")
        return

    print(f"Source 'As of Date': {as_of_date_str}")
    print("=" * 80)

    # Read source data
    print("\nReading source file: ETF_Effective_Duration.xlsx")
    source_data = read_source_data()
    if not source_data:
        return

    print(f"Total ETFs read: {len(source_data)}")

    # Create or open target file
    target_wb, target_ws = create_or_open_target_file(output_filename)

    # =============================================================================
    # STEP 1: CHECK AND UPDATE HISTORICAL DATA (As of Date from source)
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 1: CHECKING HISTORICAL DATA")
    print("=" * 80)
    print(f"Checking for historical date: {as_of_date_str}")

    historical_row = find_row_by_date(target_ws, as_of_date_str)

    if historical_row:
        print(f"  Found historical row {historical_row} with date {as_of_date_str}")

        # Get previous row data for zero handling
        hist_previous_data = get_previous_row_data(target_ws, historical_row)

        # Create column mapping
        print("  Creating column mapping for historical data...")
        hist_column_mapping, hist_zero_count = create_column_mapping(
            source_data, hist_previous_data, "historical"
        )

        # Compare and detect changes
        hist_changes = compare_and_report_changes(
            target_ws, historical_row, hist_column_mapping, f"row {historical_row}"
        )

        if hist_changes:
            print(f"  [ACTION] Correcting historical data in row {historical_row}...")
            write_data_to_row(target_ws, historical_row, as_of_date_str, hist_column_mapping)
            print(f"  [SUCCESS] Historical data CORRECTED for {as_of_date_str}")
        else:
            print(f"  [INFO] Historical data is already correct for {as_of_date_str}")
    else:
        print(f"  No historical row found for {as_of_date_str}")
        print(f"  [INFO] Will only map to yesterday's date")

    # =============================================================================
    # STEP 2: MAP TO LAST BUSINESS DATE
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 2: MAPPING TO LAST BUSINESS DATE")
    print("=" * 80)
    print(f"Checking for target date: {target_date_str}")

    existing_row = find_row_by_date(target_ws, target_date_str)

    if existing_row:
        print(f"  Found existing row {existing_row} with target date - will UPDATE")
        target_row = existing_row
        is_update = True
    else:
        print("  No existing data for target date - will CREATE NEW ROW")
        # Find next empty row
        target_row = 3
        while target_ws.cell(row=target_row, column=1).value is not None:
            target_row += 1
        print(f"  New data will be written to row {target_row}")
        is_update = False

    # Get previous row data (for handling 0 values)
    previous_data = get_previous_row_data(target_ws, target_row)

    # Create column mapping
    print("  Creating column mapping for target date...")
    column_mapping, zero_value_count = create_column_mapping(
        source_data, previous_data, "business_date"
    )

    if zero_value_count > 0:
        print(f"  [INFO] Handled {zero_value_count} zero/invalid values with previous data")

    # If updating, compare with existing data
    if is_update:
        changes_detected = compare_and_report_changes(
            target_ws, target_row, column_mapping, f"row {target_row}"
        )
        if changes_detected:
            print(f"  [ACTION] Updating row {target_row} with new data...")
        else:
            print(f"  [ACTION] Confirming row {target_row}...")
    else:
        print(f"  [ACTION] Writing new data to row {target_row}...")

    # Write data to target date row
    write_data_to_row(target_ws, target_row, target_date_str, column_mapping)

    # Display what was written
    print("\n  Data written:")
    for col_idx, data in column_mapping.items():
        etf_name = HEADER_ROW_2[col_idx - 1]
        print(f"    {openpyxl.utils.get_column_letter(col_idx)}: {data['duration']} - {etf_name}")

    # Save the target file
    print(f"\nSaving file: {output_filename}")
    target_wb.save(output_filename)
    target_wb.close()

    # Summary
    print("\n" + "=" * 80)
    print("MAPPING COMPLETED SUCCESSFULLY!")
    print("=" * 80)
    if historical_row:
        print(f"Historical data: {'CORRECTED' if hist_changes else 'VERIFIED'} for {as_of_date_str} (row {historical_row})")
    print(f"Business date: {'UPDATED' if is_update else 'CREATED'} for {target_date_str} (row {target_row})")
    print(f"Total ETFs mapped: {len(column_mapping)}")
    print(f"Zero/Invalid values handled: {zero_value_count}")
    print(f"Output file: {output_filename}")
    print("=" * 80)

if __name__ == "__main__":
    try:
        create_mapping()
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
